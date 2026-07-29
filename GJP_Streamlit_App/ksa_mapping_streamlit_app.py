from __future__ import annotations

from collections import OrderedDict
from copy import copy
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
import re
from typing import BinaryIO
from xml.etree import ElementTree as ET
import colorsys

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

CLONE_SOURCE_BY_LEVEL = {
    "P-1": "P-2",
    "G-6": "P-2",
    "G-7": "P-2",
    "P-5": "D-2",
    "P-4": "D-1",
    "P-3": "P-3",
    "P-2": "P-2",
    "D-2": "D-2",
    "D-1": "D-1",
}

MAX_LEVEL_LABEL_LEN = 15
INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")

EXTRA_THEME_PALETTE = [
    "9DC3E6",
    "A9D18E",
    "B4A7D6",
    "F4B183",
    "76D7C4",
    "F7A6A6",
    "8EAADB",
    "C5E0B3",
    "D5A6BD",
    "F8CBAD",
    "A2D2FF",
    "B9E3C6",
    "C9B6E4",
    "FDBF8F",
    "8DD3C7",
    "F4A3B4",
    "A6C8E8",
    "BFD99F",
    "CAB2D6",
    "FDBF6F",
    "80CED7",
    "F6B3C4",
    "B3C7F7",
    "B7D7A8",
    "D7B5A6",
    "A9CCE3",
    "A3E4D7",
    "D2B4DE",
    "F5CBA7",
    "AED6F1",
]


@dataclass
class ResponsibilityRow:
    original_row: int
    source_number: object
    area: str
    grouping: str
    responsibility: str
    ksa_marks: list[bool]
    output_number: int | None = None


@dataclass
class MatrixData:
    sheet_name: str
    header_row: int
    rows: list[ResponsibilityRow]
    ksa_headers: list[str]
    themes: list[str]


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def is_x_mark(value: object) -> bool:
    return clean_text(value).upper() == "X"


def responsibility_is_mandatory(cell) -> bool:
    return cell.font is not None and cell.font.bold is True


def find_header_column(headers: list[str], names: tuple[str, ...]) -> int:
    normalized = {name.lower(): idx for idx, name in enumerate(headers)}
    for name in names:
        if name.lower() in normalized:
            return normalized[name.lower()]
    for idx, header in enumerate(headers):
        h = header.lower()
        if any(name.lower() in h for name in names):
            return idx
    raise ValueError(f"Cannot find required column: {', '.join(names)}")


def is_area_header(value: str) -> bool:
    return value.lower() in {"areas of work", "area of work"}


def is_responsibility_header(value: str) -> bool:
    return "responsibilit" in value.lower()


def find_matrix_header_row(ws, max_scan_rows: int = 20) -> int | None:
    for row in range(1, min(ws.max_row, max_scan_rows) + 1):
        headers = [clean_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        if any(is_area_header(header) for header in headers) and any(is_responsibility_header(header) for header in headers):
            return row
    return None


def is_matrix_sheet(ws) -> bool:
    return find_matrix_header_row(ws) is not None


def get_matrix_sheet_names(wb) -> list[str]:
    return [ws.title for ws in wb.worksheets if is_matrix_sheet(ws)]


def extract_matrix_data(wb, sheet_name: str) -> MatrixData:
    ws = wb[sheet_name]
    header_row = find_matrix_header_row(ws)
    if header_row is None:
        raise ValueError(f"{sheet_name}: could not find a header row with Areas of work and Responsibilities.")

    headers = [clean_text(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)]
    area_idx = find_header_column(headers, ("Areas of work", "Area of work"))
    resp_idx = find_header_column(headers, ("Responsibilities", "Responsibility"))
    count_idx = None
    try:
        count_idx = find_header_column(headers, ("KSA count",))
    except ValueError:
        count_idx = None

    first_ksa_idx = (count_idx + 1) if count_idx is not None else max(area_idx, resp_idx) + 1
    ksa_headers = [clean_text(header) for header in headers[first_ksa_idx:] if clean_text(header)]
    if not ksa_headers:
        raise ValueError(f"{sheet_name}: no KSA headers were found.")

    rows: list[ResponsibilityRow] = []
    themes: OrderedDict[str, None] = OrderedDict()
    current_area = ""

    for row_num in range(header_row + 1, ws.max_row + 1):
        area = clean_text(ws.cell(row_num, area_idx + 1).value)
        if area:
            current_area = area
        responsibility = clean_text(ws.cell(row_num, resp_idx + 1).value)
        if not responsibility:
            continue
        if not current_area:
            current_area = "Unassigned"

        resp_cell = ws.cell(row_num, resp_idx + 1)
        grouping = "Mandatory" if responsibility_is_mandatory(resp_cell) else "Optional"
        source_number = ws.cell(row_num, 1).value
        marks = [
            is_x_mark(ws.cell(row_num, col).value)
            for col in range(first_ksa_idx + 1, first_ksa_idx + 1 + len(ksa_headers))
        ]

        rows.append(
            ResponsibilityRow(
                original_row=row_num,
                source_number=source_number,
                area=current_area,
                grouping=grouping,
                responsibility=responsibility,
                ksa_marks=marks,
            )
        )
        themes.setdefault(current_area, None)

    if not rows:
        raise ValueError(f"{sheet_name}: no responsibilities were found.")
    return MatrixData(sheet_name=sheet_name, header_row=header_row, rows=rows, ksa_headers=ksa_headers, themes=list(themes.keys()))


def canonical_grade_from_text(text: str) -> str | None:
    match = re.search(r"(?i)(?:^|[^A-Z0-9])([PDG])\s*-?\s*(\d)(?:$|[^A-Z0-9])", text)
    if not match:
        return None
    return f"{match.group(1).upper()}-{match.group(2)}"


def sanitize_level_label(value: str) -> str:
    label = INVALID_SHEET_CHARS.sub(" ", clean_text(value))
    label = re.sub(r"\s+", " ", label).strip(" '")
    return label or "Sheet"


def shorten_level_label(label: str, max_len: int = MAX_LEVEL_LABEL_LEN) -> str:
    label = sanitize_level_label(label)
    if len(label) <= max_len:
        return label
    words = label.split()
    if len(words) > 1:
        compact = " ".join(word[:4] for word in words)
        if len(compact) <= max_len:
            return compact
    return label[:max_len].rstrip()


def output_level_base_from_sheet_name(sheet_name: str) -> str:
    label = sanitize_level_label(sheet_name)
    if re.fullmatch(r"(?i)G\s*-?\s*\d", label):
        return label.upper().replace("-", "")

    grade = canonical_grade_from_text(label)
    if grade and len(label.split()) <= 2 and "KSA" in label.upper():
        return grade.replace("-", "")
    return shorten_level_label(label)


def make_unique_level_label(base: str, used: set[str]) -> str:
    base = shorten_level_label(base)
    candidate = base
    suffix_num = 2
    while candidate in used:
        suffix = f" {suffix_num}"
        candidate = f"{base[: MAX_LEVEL_LABEL_LEN - len(suffix)].rstrip()}{suffix}"
        suffix_num += 1
    used.add(candidate)
    return candidate


def build_sheet_mapping(sheet_names: list[str]) -> dict[str, str]:
    used: set[str] = set()
    return {
        sheet_name: make_unique_level_label(output_level_base_from_sheet_name(sheet_name), used)
        for sheet_name in sheet_names
    }


def workbook_from_path_or_upload(path: Path | None = None, upload: BinaryIO | None = None, data_only: bool = False):
    if upload is not None:
        upload.seek(0)
        return load_workbook(BytesIO(upload.read()), data_only=data_only)
    if path is None:
        raise ValueError("A workbook path or upload is required.")
    return load_workbook(path, data_only=data_only)


def workbook_bytes_from_path_or_upload(path: Path | None = None, upload: BinaryIO | None = None) -> bytes:
    if upload is not None:
        upload.seek(0)
        return upload.read()
    if path is None:
        raise ValueError("A workbook path or upload is required.")
    return path.read_bytes()


def get_theme_rgb_map(wb) -> dict[int, str]:
    if not wb.loaded_theme:
        return {}
    root = ET.fromstring(wb.loaded_theme)
    ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    scheme = root.find(".//a:clrScheme", ns)
    if scheme is None:
        return {}

    order = ["lt1", "dk1", "lt2", "dk2", "accent1", "accent2", "accent3", "accent4", "accent5", "accent6", "hlink", "folHlink"]
    colors: dict[int, str] = {}
    for idx, name in enumerate(order):
        node = scheme.find(f"a:{name}", ns)
        if node is None or len(node) == 0:
            continue
        color_node = node[0]
        value = color_node.attrib.get("val") or color_node.attrib.get("lastClr")
        if value:
            colors[idx] = value.upper()
    return colors


def apply_tint(rgb: str, tint: float) -> str:
    rgb = rgb[-6:]
    parts = [int(rgb[i : i + 2], 16) for i in (0, 2, 4)]
    tinted = []
    for part in parts:
        if tint < 0:
            value = int(round(part * (1.0 + tint)))
        else:
            value = int(round(part + (255 - part) * tint))
        tinted.append(max(0, min(255, value)))
    return "".join(f"{part:02X}" for part in tinted)


def fill_to_rgb(wb, fill: PatternFill) -> str:
    color = fill.fgColor
    if color.type == "rgb" and color.rgb:
        return color.rgb[-6:]
    if color.type == "theme" and color.theme is not None:
        base = get_theme_rgb_map(wb).get(color.theme, "FFFFFF")
        return apply_tint(base, color.tint or 0.0)
    if color.type == "indexed" and color.indexed is not None:
        return "FFFFFF"
    return "FFFFFF"


def get_theme_fills(wb) -> list[PatternFill]:
    ws = wb["1. Theme_Names"]
    fills = []
    for row in range(7, 17):
        fill = copy(ws.cell(row, 2).fill)
        if fill.fill_type:
            fills.append(fill)
        else:
            fills.append(PatternFill(fill_type="solid", fgColor="FFFFFFFF"))
    return fills


def rgb_parts(rgb: str) -> tuple[int, int, int]:
    rgb = rgb.upper()[-6:]
    return tuple(int(rgb[i : i + 2], 16) for i in (0, 2, 4))


def rgb_distance(first: str, second: str) -> float:
    red_a, green_a, blue_a = rgb_parts(first)
    red_b, green_b, blue_b = rgb_parts(second)
    return ((red_a - red_b) ** 2 + (green_a - green_b) ** 2 + (blue_a - blue_b) ** 2) ** 0.5


def color_is_distinct(rgb: str, existing: set[str], min_distance: float) -> bool:
    return all(rgb_distance(rgb, other) >= min_distance for other in existing)


def unique_palette_colors(existing_rgb: set[str], needed: int) -> list[str]:
    colors: list[str] = []
    seen = {rgb.upper()[-6:] for rgb in existing_rgb}

    def add_color(rgb: str, min_distance: float) -> bool:
        rgb = rgb.upper()[-6:]
        if rgb in seen:
            return False
        if not color_is_distinct(rgb, seen, min_distance):
            return False
        colors.append(rgb)
        seen.add(rgb)
        return True

    for min_distance in (74, 64, 54, 44):
        for rgb in EXTRA_THEME_PALETTE:
            if len(colors) >= needed:
                return colors
            add_color(rgb, min_distance)

    hue_index = 0
    min_distance = 58
    attempts_at_distance = 0
    while len(colors) < needed:
        hue = (hue_index * 0.618033988749895) % 1.0
        saturation = 0.24 + 0.10 * (hue_index % 4) / 3
        value = 0.96
        red, green, blue = colorsys.hsv_to_rgb(hue, saturation, value)
        rgb = f"{int(red * 255):02X}{int(green * 255):02X}{int(blue * 255):02X}"
        if add_color(rgb, min_distance):
            attempts_at_distance = 0
        else:
            attempts_at_distance += 1
            if attempts_at_distance > 720 and min_distance > 0:
                min_distance = max(0, min_distance - 8)
                attempts_at_distance = 0
        hue_index += 1
    return colors


def make_solid_fill(rgb: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=f"FF{rgb.upper()[-6:]}")


def copy_style(src, dst) -> None:
    if src.has_style:
        dst._style = copy(src._style)
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)


def set_bold(cell, bold: bool) -> None:
    font = copy(cell.font)
    font.bold = bold
    cell.font = font


def clear_values(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row, col).value = None


def ensure_sheet_capacity(ws, needed_rows: int) -> None:
    if ws.max_row < needed_rows:
        ws.insert_rows(ws.max_row + 1, amount=needed_rows - ws.max_row)


def ensure_column_capacity(ws, needed_cols: int) -> None:
    if ws.max_column < needed_cols:
        ws.insert_cols(ws.max_column + 1, amount=needed_cols - ws.max_column)


def write_theme_names(wb, themes: list[str]) -> dict[str, PatternFill]:
    ws = wb["1. Theme_Names"]
    fills = get_theme_fills(wb)
    if len(themes) > 10:
        extra_count = len(themes) - 10
        ws.insert_rows(17, amount=extra_count)
        for row in range(17, 17 + extra_count):
            for col in range(1, 3):
                copy_style(ws.cell(16, col), ws.cell(row, col))

        existing_rgb = {fill_to_rgb(wb, fill) for fill in fills}
        extra_colors = unique_palette_colors(existing_rgb, extra_count)
        fills.extend(make_solid_fill(rgb) for rgb in extra_colors)

    clear_values(ws, 7, 6 + max(len(themes), 10), 1, 2)
    for idx, theme in enumerate(themes):
        cell = ws.cell(7 + idx, 2)
        ws.cell(7 + idx, 1).value = idx + 1
        cell.value = theme
        cell.fill = copy(fills[idx])
    for idx in range(len(themes), len(fills)):
        ws.cell(7 + idx, 1).value = idx + 1
        ws.cell(7 + idx, 2).fill = copy(fills[idx])
    return {theme: fills[idx] for idx, theme in enumerate(themes)}


def rows_for_output(matrix: MatrixData) -> list[ResponsibilityRow]:
    rows = sorted(matrix.rows, key=lambda r: r.original_row)
    for idx, row in enumerate(rows, start=1):
        row.output_number = idx
    return rows


def sheet_names_for_level(level: str) -> tuple[str, str, str]:
    level = shorten_level_label(level)
    return f"Input_{level}_KSA_Check", f"Output_{level}_Resp", f"Output_{level}_KSA"


def level_from_template_sheet_name(sheet_name: str) -> str | None:
    patterns = [
        r"^Input_(.+)_KSA_Check$",
        r"^Output_(.+)_Resp$",
        r"^Output_(.+)_KSA$",
    ]
    for pattern in patterns:
        match = re.match(pattern, sheet_name)
        if match:
            return match.group(1)
    return None


def copy_template_sheet(wb, source_title: str, target_title: str) -> None:
    if len(target_title) > 31:
        raise ValueError(f"Generated sheet name is too long for Excel: {target_title}")
    if target_title in wb.sheetnames:
        return
    if source_title not in wb.sheetnames:
        raise ValueError(f"Template sheet is missing: {source_title}")
    copied = wb.copy_worksheet(wb[source_title])
    copied.title = target_title


def choose_clone_level(wb, level: str) -> str:
    canonical_level = canonical_grade_from_text(level) or level
    preferred = CLONE_SOURCE_BY_LEVEL.get(canonical_level)
    if preferred and all(name in wb.sheetnames for name in sheet_names_for_level(preferred)):
        return preferred
    if all(name in wb.sheetnames for name in sheet_names_for_level(canonical_level)):
        return canonical_level
    if all(name in wb.sheetnames for name in sheet_names_for_level(level)):
        return shorten_level_label(level)
    fallback_order = ["P-3", "P-2", "D-1", "D-2"]
    for fallback in fallback_order:
        if all(name in wb.sheetnames for name in sheet_names_for_level(fallback)):
            return fallback
    raise ValueError("No complete Input/Output sheet set is available to clone.")


def ensure_level_sheets(wb, levels: list[str]) -> None:
    for level in levels:
        input_sheet, resp_sheet, ksa_sheet = sheet_names_for_level(level)
        clone_level = choose_clone_level(wb, level)
        clone_input, clone_resp, clone_ksa = sheet_names_for_level(clone_level)

        if input_sheet not in wb.sheetnames:
            copy_template_sheet(wb, clone_input, input_sheet)
        if resp_sheet not in wb.sheetnames:
            copy_template_sheet(wb, clone_resp, resp_sheet)
        if ksa_sheet not in wb.sheetnames:
            copy_template_sheet(wb, clone_ksa, ksa_sheet)


def remove_unused_level_sheets(wb, levels_to_keep: set[str]) -> None:
    for sheet_name in list(wb.sheetnames):
        level = level_from_template_sheet_name(sheet_name)
        if level is not None and level not in levels_to_keep:
            del wb[sheet_name]


def reorder_level_sheets(wb, levels: list[str]) -> None:
    non_level_sheets = [ws for ws in wb.worksheets if level_from_template_sheet_name(ws.title) is None]
    ordered_level_sheets = []
    for level in levels:
        for title in sheet_names_for_level(level):
            if title in wb.sheetnames:
                ordered_level_sheets.append(wb[title])
    wb._sheets = non_level_sheets + ordered_level_sheets


def write_input_sheet(wb, level: str, matrix: MatrixData) -> None:
    input_sheet, _, _ = sheet_names_for_level(level)
    ws = wb[input_sheet]
    max_ksa = len(matrix.ksa_headers)
    needed_rows = 9 + len(matrix.rows)
    needed_cols = 4 + max_ksa
    ensure_sheet_capacity(ws, needed_rows)
    ensure_column_capacity(ws, needed_cols)
    clear_values(ws, 10, max(ws.max_row, needed_rows + 10), 1, max(ws.max_column, needed_cols))

    for col, header in enumerate(matrix.ksa_headers[:max_ksa], start=5):
        copy_style(ws.cell(9, 5), ws.cell(9, col))
        ws.cell(9, col).value = header

    for idx, row_data in enumerate(matrix.rows, start=10):
        template_row = 10
        for col in range(1, 5 + max_ksa):
            style_col = col if col <= 4 else 5
            copy_style(ws.cell(template_row, style_col), ws.cell(idx, col))
        ws.cell(idx, 1).value = idx - 9
        ws.cell(idx, 2).value = row_data.grouping
        ws.cell(idx, 3).value = row_data.area
        ws.cell(idx, 4).value = row_data.responsibility
        set_bold(ws.cell(idx, 2), False)
        set_bold(ws.cell(idx, 3), False)
        set_bold(ws.cell(idx, 4), row_data.grouping == "Mandatory")
        for offset, marked in enumerate(row_data.ksa_marks[:max_ksa], start=5):
            ws.cell(idx, offset).value = "x" if marked else None


def write_resp_output(wb, level: str, rows: list[ResponsibilityRow], theme_fills: dict[str, PatternFill]) -> None:
    _, resp_sheet, _ = sheet_names_for_level(level)
    ws = wb[resp_sheet]
    ensure_sheet_capacity(ws, 8 + len(rows))
    clear_values(ws, 9, max(ws.max_row, 9 + len(rows) + 10), 1, 3)

    for out_row, row_data in enumerate(rows, start=9):
        for col in range(1, 4):
            copy_style(ws.cell(9, col), ws.cell(out_row, col))
        area_fill = copy(theme_fills.get(row_data.area, PatternFill()))
        ws.cell(out_row, 1).fill = copy(area_fill)
        ws.cell(out_row, 2).fill = copy(area_fill)
        ws.cell(out_row, 3).fill = PatternFill(fill_type=None)
        ws.cell(out_row, 1).value = row_data.area
        ws.cell(out_row, 2).value = row_data.output_number
        ws.cell(out_row, 3).value = row_data.responsibility
        is_mandatory = row_data.grouping == "Mandatory"
        set_bold(ws.cell(out_row, 1), is_mandatory)
        set_bold(ws.cell(out_row, 2), is_mandatory)
        set_bold(ws.cell(out_row, 3), is_mandatory)


def build_ksa_output(matrix: MatrixData, rows: list[ResponsibilityRow]) -> list[tuple[str, list[int]]]:
    result: list[tuple[str, list[int]]] = []
    for idx, header in enumerate(matrix.ksa_headers):
        linked_numbers = [
            row.output_number
            for row in rows
            if row.output_number is not None and idx < len(row.ksa_marks) and row.ksa_marks[idx]
        ]
        if linked_numbers:
            result.append((header, linked_numbers))
    return result


def write_ksa_output(
    wb,
    level: str,
    matrix: MatrixData,
    rows: list[ResponsibilityRow],
    theme_fills: dict[str, PatternFill],
) -> None:
    responsibilities_per_line = 10
    _, _, ksa_sheet = sheet_names_for_level(level)
    ws = wb[ksa_sheet]
    ksa_rows = build_ksa_output(matrix, rows)
    rendered_row_count = sum(
        max(1, (len(numbers) + responsibilities_per_line - 1) // responsibilities_per_line)
        for _, numbers in ksa_rows
    )
    ensure_sheet_capacity(ws, 10 + rendered_row_count)
    max_links = max((len(numbers) for _, numbers in ksa_rows), default=0)
    max_resp_cols = max(min(max_links, responsibilities_per_line), 1)
    ensure_column_capacity(ws, 3 + max_resp_cols)
    clear_values(ws, 11, max(ws.max_row, 11 + rendered_row_count + 10), 2, max(ws.max_column, 3 + max_resp_cols))
    if hasattr(ws.conditional_formatting, "_cf_rules"):
        ws.conditional_formatting._cf_rules.clear()

    number_to_theme = {row.output_number: row.area for row in rows}
    number_to_mandatory = {row.output_number: row.grouping == "Mandatory" for row in rows}
    last_output_row = max(ws.max_row, 11 + rendered_row_count + 10)
    for row_idx in range(11, last_output_row + 1):
        for col_idx in range(4, 4 + max_resp_cols):
            ws.cell(row_idx, col_idx).fill = PatternFill(fill_type=None)
            set_bold(ws.cell(row_idx, col_idx), False)

    output_row = 11
    for ksa_number, (ksa, numbers) in enumerate(ksa_rows, start=1):
        number_lines = [
            numbers[start : start + responsibilities_per_line]
            for start in range(0, len(numbers), responsibilities_per_line)
        ] or [[]]
        first_row = output_row

        for number_line in number_lines:
            for col in range(2, 4 + max_resp_cols):
                style_col = col if col <= 3 else 4
                copy_style(ws.cell(11, style_col), ws.cell(output_row, col))
                ws.cell(output_row, col).value = None
                if col >= 4:
                    ws.cell(output_row, col).fill = PatternFill(fill_type=None)
                    set_bold(ws.cell(output_row, col), False)

            for offset, number in enumerate(number_line, start=4):
                cell = ws.cell(output_row, offset)
                cell.value = number
                cell.fill = copy(theme_fills.get(number_to_theme.get(number, ""), PatternFill()))
                set_bold(cell, number_to_mandatory.get(number, False))
            output_row += 1

        last_ksa_row = output_row - 1
        ws.cell(first_row, 2).value = ksa_number
        ws.cell(first_row, 3).value = ksa
        if last_ksa_row > first_row:
            ws.merge_cells(start_row=first_row, start_column=2, end_row=last_ksa_row, end_column=2)
            ws.merge_cells(start_row=first_row, start_column=3, end_row=last_ksa_row, end_column=3)


def fill_template_workbook(template_bytes: bytes, source_bytes: bytes, mapping: dict[str, str]) -> tuple[bytes, dict[str, object]]:
    template_wb = load_workbook(BytesIO(template_bytes), data_only=False)
    source_wb = load_workbook(BytesIO(source_bytes), data_only=False)

    matrices = {sheet: extract_matrix_data(source_wb, sheet) for sheet in mapping}
    levels = list(mapping.values())
    duplicate_levels = sorted({level for level in levels if levels.count(level) > 1})
    if duplicate_levels:
        raise ValueError(f"More than one source sheet maps to the same level: {', '.join(duplicate_levels)}")

    ensure_level_sheets(template_wb, levels)
    remove_unused_level_sheets(template_wb, set(levels))
    reorder_level_sheets(template_wb, levels)

    themes: OrderedDict[str, None] = OrderedDict()
    for matrix in matrices.values():
        for theme in matrix.themes:
            themes.setdefault(theme, None)
    theme_list = list(themes.keys())

    theme_fills = write_theme_names(template_wb, theme_list)
    for source_sheet, level in mapping.items():
        matrix = matrices[source_sheet]
        rows = rows_for_output(matrix)
        write_input_sheet(template_wb, level, matrix)
        write_resp_output(template_wb, level, rows, theme_fills)
        write_ksa_output(template_wb, level, matrix, rows, theme_fills)

    output = BytesIO()
    template_wb.save(output)
    summary = {
        "themes": theme_list,
        "processed": {
            source_sheet: {
                "level": level,
                "responsibilities": len(matrices[source_sheet].rows),
                "ksas": len(matrices[source_sheet].ksa_headers),
            }
            for source_sheet, level in mapping.items()
        },
    }
    return output.getvalue(), summary


def render_ksa_conversion_tables() -> None:
    import streamlit as st

    st.title("KSA Conversion Tables")
    st.caption("Upload both Excel files. They are processed only for the current browser session.")

    source_upload = st.file_uploader(
        "1) Upload the Excel workbook to convert",
        type=["xlsx"],
        key="ksa_source_workbook",
        help="Upload the workbook containing the Areas of work and Responsibilities matrix.",
    )
    template_upload = st.file_uploader(
        "2) Upload template: 6.2 Transposing KSA mapping_Theme Version_Investigation.xlsx",
        type=["xlsx"],
        key="ksa_template_workbook",
        help="This template supplies the output worksheets, formulas, colors, and formatting.",
    )

    if source_upload is None or template_upload is None:
        missing = []
        if source_upload is None:
            missing.append("the Excel workbook to convert")
        if template_upload is None:
            missing.append("the 6.2 Transposing KSA mapping template")
        st.info(f"Please upload {' and '.join(missing)} to continue.")
        return

    try:
        source_bytes = source_upload.getvalue()
        template_bytes = template_upload.getvalue()
        source_wb = load_workbook(BytesIO(source_bytes), data_only=False)
        matrix_sheets = get_matrix_sheet_names(source_wb)
        if not matrix_sheets:
            st.error("No matrix sheets were found. The sheet must contain `Areas of work` and `Responsibilities` headers.")
            return
    except Exception as exc:
        st.error(f"Could not read the uploaded matrix workbook: {exc}")
        return

    st.subheader("Detected levels")
    mapping = build_sheet_mapping(matrix_sheets)

    st.dataframe(
        [{"Source sheet": sheet, "Output level": level} for sheet, level in mapping.items()],
        use_container_width=True,
        hide_index=True,
    )
    st.caption("Each source sheet gets its own Input/Output sheet group. Output names are shortened only when Excel's 31-character sheet-name limit requires it.")

    st.subheader("Preview")
    preview_cols = st.columns(min(4, len(matrix_sheets)))
    for idx, sheet in enumerate(matrix_sheets[:4]):
        matrix = extract_matrix_data(source_wb, sheet)
        with preview_cols[idx % len(preview_cols)]:
            st.metric(sheet, f"{len(matrix.rows)} responsibilities", f"{len(matrix.ksa_headers)} KSAs")
    if len(matrix_sheets) > 4:
        st.caption(f"{len(matrix_sheets) - 4} additional sheet(s) will also be converted.")

    if st.button("Generate filled workbook", type="primary"):
        try:
            filled_bytes, summary = fill_template_workbook(template_bytes, source_bytes, mapping)
        except Exception as exc:
            st.error(f"Generation failed: {exc}")
            return

        st.success("Workbook generated.")
        st.write("Themes:", ", ".join(summary["themes"]))
        for sheet, info in summary["processed"].items():
            st.write(
                f"{sheet} -> {info['level']}: "
                f"{info['responsibilities']} responsibilities, {info['ksas']} KSAs"
            )
        st.download_button(
            "Download filled workbook",
            data=filled_bytes,
            file_name="KSA_mapping_filled_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def run_app() -> None:
    import streamlit as st

    st.set_page_config(page_title="KSA Mapping Generator", page_icon="📊", layout="wide")
    render_ksa_conversion_tables()


if __name__ == "__main__":
    run_app()
